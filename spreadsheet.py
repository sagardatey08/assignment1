import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

current_direct=os.path.dirname(__file__)
filename=os.path.join(current_direct,"matrix.xlsx")

wb=openpyxl.Workbook()
print("Sheets:",wb.sheetnames)
sheet=wb.active
sheet.title="Hiten"
print("Sheets title:",sheet.title)
"""A=column_index_from_string('A')
Z=column_index_from_string('Z')"""
for i in sheet['A1':'Z100']:
    for j in i:
        #j1=get_column_letter(j)
        
        #print(j1,end=" ")
        j.value=j.coordinate
        print(j.value,end="")
        
        #print(sheet.cell(row=i,column=j).value,end=" ")

    print()


wb.save(filename)
